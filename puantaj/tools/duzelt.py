#!/usr/bin/env python3
"""
PUANTAJ CETVELİ ONARIM ARACI
============================

Mevcut puantaj Excel dosyanızdaki üç yapısal hatayı düzeltir, düzeni bozmadan:

  1. YANLIŞ TOPLAM FORMÜLLERİ
     Sayfalar birbirinden kopyalandığı için COUNTIF aralıkları ayın gün
     sayısıyla uyuşmuyor. Örneğin 31 günlük bir ayda izin sayımı D:AG
     (30 gün) yapılıyor, ayın 31'i hiç sayılmıyor.

  2. BÜYÜK/KÜÇÜK HARF TUTARSIZLIĞI
     Hücrelere kimi zaman "ç" / "i", kimi zaman "Ç" / "İ" yazılmış.
     COUNTIF(...,"İ") küçük "i" yazılmış günleri saymayabiliyor; bu da
     izin günlerinin eksik görünmesine yol açıyor.

  3. BAŞLIKTAKİ YANLIŞ TARİH
     Sayfa kopyalanırken başlık hücresindeki tarih güncellenmemiş.

Ayrıca dosyayı kullanışlı hale getirir: kodlara renk (koşullu biçimlendirme),
gün hücrelerine açılır liste, başlık satırı dondurma, hafta sonu vurgusu.

KULLANIM
    python3 duzelt.py PUANTAJ_CETVELI.xlsx
    python3 duzelt.py PUANTAJ_CETVELI.xlsx -o CIKTI.xlsx

Gereksinim:  pip install openpyxl
"""

import argparse
import calendar
import datetime as dt
import re
import sys
from collections import Counter

try:
    import openpyxl
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl kurulu değil. Kurmak için:  pip install openpyxl")


AYLAR = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
         "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
AY_NO = {ad: i + 1 for i, ad in enumerate(AYLAR)}

# Toplam sütunu başlığı -> sayılacak kod
TOPLAM_BASLIK = {
    "ÇALIŞILAN GÜN": "Ç",
    "İZİNLİ GÜN": "İ",
    "RAPORLU GÜN": "R",
    "YARIM GÜN": "Y",
    "DEVAMSIZ": "D",
    "DEVAMSIZ GÜN": "D",
    "ÜCRETSİZ İZİN": "Ü",
}
# Renklendirilecek kodlar: kod -> (yazı rengi, dolgu rengi)
KOD_RENK = {
    "Ç": ("1C7A4A", "DFF3E7"),
    "İ": ("B06A00", "FDEFD6"),
    "R": ("A32020", "FBE0E0"),
    "Y": ("5A4BB5", "E8E5FA"),
    "D": ("4A5464", "E3E7ED"),
    "Ü": ("8A6D3B", "F2EAD9"),
    "Dİ": ("0B6F86", "DCF0F5"),
}
BASLIK_KELIMELERI = {
    "ADI - SOYADI", "ADI SOYADI", "AD SOYAD", "TC NO", "ÇALIŞMA YERİ",
    "PUANTAJ CETVELİ", "MESAİ SAATI", "FAZLA MESAİ SAATİ", "FAZLA MESAİ TUTARI",
    "TOPLAM", *TOPLAM_BASLIK,
}
HAFTA_SONU_DOLGU = PatternFill("solid", fgColor="E8ECF3")


def tr_buyuk(s):
    """Türkçe'ye duyarlı büyük harf: i -> İ, ı -> I."""
    return str(s).replace("i", "İ").replace("ı", "I").upper()


def baslik_mi(deger):
    return tr_buyuk(str(deger or "").strip()) in BASLIK_KELIMELERI


def sayfadan_ay(sayfa_adi):
    """'TEMMUZ 2025' -> (2025, 7). Çözülemezse None."""
    t = re.sub(r"\s+", " ", tr_buyuk(sayfa_adi)).strip()
    m = re.match(r"^([A-ZÇĞİÖŞÜ]+)\s+(\d{4})$", t)
    if m and m.group(1) in AY_NO:
        return int(m.group(2)), AY_NO[m.group(1)]
    return None


def kod_normalize(ham):
    """Hücre içeriğini standart koda çevirir. (yeni_deger, degisti_mi)"""
    if ham is None or not isinstance(ham, str):
        return ham, False
    s = ham.strip()
    if not s:
        return (None, True) if ham != "" else (ham, False)
    b = tr_buyuk(s)
    if b == "DI":          # "di" / "Di" -> doğum izni
        b = "Dİ"
    if b == "C":           # klavyeden Ç yerine C
        b = "Ç"
    return (b, True) if b != ham else (ham, False)


class Rapor:
    def __init__(self):
        self.formul = 0
        self.formul_sayfa = set()
        self.kod = Counter()
        self.tarih = []
        self.bilinmeyen = Counter()
        self.sayfa = 0
        self.atlanan = []

    def yaz(self):
        c = print
        c("")
        c("=" * 66)
        c("  ONARIM RAPORU")
        c("=" * 66)
        c(f"  İşlenen sayfa                 : {self.sayfa}")
        if self.atlanan:
            c(f"  Atlanan sayfa (ay adı yok)    : {len(self.atlanan)}  "
              f"{', '.join(self.atlanan[:5])}{'…' if len(self.atlanan) > 5 else ''}")
        c("")
        c(f"  1) Düzeltilen toplam formülü  : {self.formul:>6}  "
          f"({len(self.formul_sayfa)} sayfada)")
        if self.formul_sayfa:
            liste = sorted(self.formul_sayfa)
            c(f"     Etkilenen sayfalar         : {', '.join(liste[:6])}"
              f"{f' … (+{len(liste) - 6})' if len(liste) > 6 else ''}")
        c("")
        toplam_kod = sum(self.kod.values())
        c(f"  2) Standartlaştırılan hücre   : {toplam_kod:>6}")
        for (eski, yeni), n in self.kod.most_common(10):
            c(f"     {eski!r:>8}  ->  {yeni!r:<8} {n:>6} hücre")
        c("")
        c(f"  3) Düzeltilen başlık tarihi   : {len(self.tarih):>6}")
        if self.tarih:
            c(f"     Örnek: {self.tarih[0]}")
        c("")
        if self.bilinmeyen:
            c("  ⚠ TANINMAYAN KODLAR (dokunulmadı, kontrol edin):")
            for k, n in self.bilinmeyen.most_common():
                c(f"     {k!r:>8} {n:>6} hücre")
            c("")
        c("=" * 66)


def bloklari_bul(ws):
    """Sayfadaki her tablo bloğu için (baslik_satiri, toplam_sutunlari) döndürür.
    toplam_sutunlari: {sutun_no: kod}"""
    bloklar = []
    for r in range(1, ws.max_row + 1):
        if tr_buyuk(str(ws.cell(r, 2).value or "").strip()) not in ("ADI - SOYADI", "ADI SOYADI", "AD SOYAD"):
            continue
        toplamlar = {}
        for c in range(4, ws.max_column + 1):
            etiket = tr_buyuk(str(ws.cell(r, c).value or "").strip())
            if etiket in TOPLAM_BASLIK:
                toplamlar[c] = TOPLAM_BASLIK[etiket]
        bloklar.append((r, toplamlar))
    return bloklar


def sayfayi_onar(ws, yil, ay, rapor, sade=False):
    gun_adedi = calendar.monthrange(yil, ay)[1]
    ilk_gun_sutun = 4                       # D sütunu = ayın 1'i
    son_gun_sutun = 3 + gun_adedi
    son_gun_harf = get_column_letter(son_gun_sutun)

    # --- 3) başlıktaki tarih ---
    for r in range(1, 6):
        for c in range(1, min(ws.max_column, 20) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, dt.datetime) and (v.year != yil or v.month != ay):
                rapor.tarih.append(f"{ws.title}: {v:%Y-%m-%d} -> {yil}-{ay:02d}-01")
                ws.cell(r, c).value = dt.datetime(yil, ay, 1)

    bloklar = bloklari_bul(ws)
    if not bloklar:
        return

    blok_sinirlari = [b[0] for b in bloklar] + [ws.max_row + 1]

    for i, (baslik_satiri, toplamlar) in enumerate(bloklar):
        ilk_veri = baslik_satiri + 1
        son_veri = blok_sinirlari[i + 1] - 1

        # --- başlık satırındaki gün numaralarını tazele ---
        for g in range(1, gun_adedi + 1):
            ws.cell(baslik_satiri, 3 + g).value = g

        for r in range(ilk_veri, son_veri + 1):
            ad = str(ws.cell(r, 2).value or "").strip()
            if not ad or baslik_mi(ad):
                continue

            # --- 2) gün hücrelerini standartlaştır ---
            for c in range(ilk_gun_sutun, son_gun_sutun + 1):
                h = ws.cell(r, c)
                eski = h.value
                yeni, degisti = kod_normalize(eski)
                if degisti:
                    h.value = yeni
                    rapor.kod[(eski, yeni)] += 1
                if isinstance(h.value, str) and h.value not in KOD_RENK:
                    rapor.bilinmeyen[h.value] += 1

            # --- 1) toplam formüllerini doğru aralıkla yeniden yaz ---
            for c, kod in toplamlar.items():
                dogru = f'=COUNTIF(D{r}:{son_gun_harf}{r},"{kod}")'
                mevcut = ws.cell(r, c).value
                if str(mevcut or "").replace(" ", "") != dogru.replace(" ", ""):
                    ws.cell(r, c).value = dogru
                    rapor.formul += 1
                    rapor.formul_sayfa.add(ws.title)

    if sade:
        return

    # ---------- kullanım kolaylıkları ----------
    ilk_baslik = bloklar[0][0]
    son_satir = max(ws.max_row, ilk_baslik + 1)
    aralik = f"D{ilk_baslik + 1}:{son_gun_harf}{son_satir}"

    # kodlara renk (Excel'de elle yazılan yeni kodlar da renklenir)
    ws.conditional_formatting = type(ws.conditional_formatting)()
    for kod, (yazi, dolgu) in KOD_RENK.items():
        ws.conditional_formatting.add(aralik, CellIsRule(
            operator="equal", formula=[f'"{kod}"'],
            font=Font(bold=True, color=yazi),
            fill=PatternFill("solid", start_color=dolgu, end_color=dolgu)))

    # açılır liste (uyarı gösterir ama yazmayı engellemez)
    dv = DataValidation(
        type="list", formula1='"' + ",".join(KOD_RENK) + '"',
        allow_blank=True, showErrorMessage=False)
    dv.promptTitle = "Puantaj kodu"
    dv.prompt = "Ç=Çalıştı  İ=İzinli  R=Raporlu  Y=Yarım gün  D=Devamsız  Ü=Ücretsiz izin  Dİ=Doğum izni"
    ws.add_data_validation(dv)
    dv.add(aralik)

    # hafta sonlarını başlık satırında vurgula + sütun genişlikleri
    for g in range(1, gun_adedi + 1):
        harf = get_column_letter(3 + g)
        ws.column_dimensions[harf].width = 4.2
        if dt.date(yil, ay, g).weekday() >= 5:
            for baslik_satiri, _ in bloklar:
                ws.cell(baslik_satiri, 3 + g).fill = HAFTA_SONU_DOLGU

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    for c in bloklar[0][1]:
        ws.column_dimensions[get_column_letter(c)].width = 12

    # başlık satırını ve isim sütunlarını dondur
    ws.freeze_panes = f"D{ilk_baslik + 1}"


def main():
    ap = argparse.ArgumentParser(
        description="Puantaj cetvelindeki formül ve kod hatalarını düzeltir.")
    ap.add_argument("girdi", help="Onarılacak .xlsx dosyası")
    ap.add_argument("-o", "--cikti", help="Çıktı dosyası (varsayılan: <girdi>_DUZELTILMIS.xlsx)")
    ap.add_argument("--sade", action="store_true",
                    help="Sadece hataları düzelt; renk/açılır liste ekleme")
    a = ap.parse_args()

    cikti = a.cikti or re.sub(r"\.xlsx$", "", a.girdi, flags=re.I) + "_DUZELTILMIS.xlsx"

    print(f"Okunuyor: {a.girdi}")
    wb = openpyxl.load_workbook(a.girdi)
    rapor = Rapor()

    for ws in wb.worksheets:
        ay_bilgi = sayfadan_ay(ws.title)
        if not ay_bilgi:
            rapor.atlanan.append(ws.title)
            continue
        sayfayi_onar(ws, ay_bilgi[0], ay_bilgi[1], rapor, sade=a.sade)
        rapor.sayfa += 1

    wb.save(cikti)
    rapor.yaz()
    print(f"\n✓ Kaydedildi: {cikti}\n")
    print("  Orijinal dosyanıza dokunulmadı. Yeni dosyayı açıp toplamları")
    print("  kontrol ettikten sonra kullanmaya başlayabilirsiniz.\n")


if __name__ == "__main__":
    main()
